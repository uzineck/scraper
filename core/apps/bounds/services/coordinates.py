from dataclasses import dataclass
from abc import ABC, abstractmethod

from openpyxl.worksheet.worksheet import Worksheet

from core.apps.bounds.entities.coordinates import StartCoordinates, EndCoordinates
from core.common.utils import Utils


@dataclass
class BaseBoundsService(ABC):
    sheet: Worksheet
    sheet_name: str

    @staticmethod
    @abstractmethod
    def find_start_coordinates(sheet: Worksheet, target_word: str) -> StartCoordinates:
        ...

    @abstractmethod
    def find_end_coordinates(self, start_row: int) -> EndCoordinates:
        ...

    @abstractmethod
    def find_sheet_bounds(self) -> tuple[StartCoordinates, EndCoordinates]:
        ...


class BoundsService(BaseBoundsService):
    @staticmethod
    def find_start_coordinates(sheet: Worksheet, target_word: str) -> StartCoordinates:
        for row_coord, row_cells in enumerate(sheet.iter_rows(values_only=True), start=1):
            for col_coord, cell_value in enumerate(row_cells, start=1):
                if cell_value == target_word:
                    return StartCoordinates(col_coord, row_coord)

        return StartCoordinates(0, 0)

    def find_sheet_bounds(self) -> tuple[StartCoordinates, EndCoordinates]:
        target_word = Utils.get_target_word(column=1, sheet_name=self.sheet_name)

        start_coordinates = self.find_start_coordinates(self.sheet, target_word)
        start_row = start_coordinates.Row
        # start_column = start_coordinates.Column

        end_coordinates = self.find_end_coordinates(start_row=start_row)
        # end_row = end_coordinates.Row
        # end_column = end_coordinates.Column

        return start_coordinates, end_coordinates

    def find_end_coordinates(self, start_row: int) -> EndCoordinates:
        end_column = self._find_end_column(start_row)
        end_row = self._find_end_row(start_row=start_row, end_column=end_column)

        return EndCoordinates(end_column, end_row)

    def _find_end_column(self,
                         start_row: int,
                         ) -> int:

        end_column_coord = None
        for row_items in self.sheet.iter_rows(min_row=start_row, max_row=start_row):
            for col_idx, cell in enumerate(row_items, start=1):
                if cell.value is None:
                    if cell.coordinate not in self.sheet.merged_cells:
                        return end_column_coord
                else:
                    end_column_coord = col_idx
        return end_column_coord

    def _find_end_row(self,
                      start_row: int,
                      end_column: int,
                      max_row: int = 200,
                      ) -> int:
        end_row_coord = None
        for row_coord in range(start_row, max_row):
            cell = self.sheet.cell(row=row_coord, column=end_column)
            if cell.value is not None:
                merged_cell = Utils.is_merged_cell(self.sheet, cell)
                if merged_cell:
                    end_row_coord = merged_cell.max_row
                else:
                    end_row_coord = row_coord
        return end_row_coord


